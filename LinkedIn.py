from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from datetime import datetime
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook


class LinkedInUnreadMonitor:
    def __init__(self, chromedriver_path):
        self.chromedriver_path = chromedriver_path
        self.driver = None

    def login(self, username, password):
        service = Service(self.chromedriver_path)
        self.driver = webdriver.Chrome(service=service)

        try:
            self.driver.get('https://www.linkedin.com')

            email_input = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, 'session_key')))
            email_input.send_keys(username)

            password_input = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, 'session_password')))
            password_input.send_keys(password)

            sign_in_button = self.driver.find_element(By.XPATH, '//button[@type="submit"]')
            sign_in_button.click()

        except Exception as e:
            print("An error occurred during login:", e)

    def get_unread_counts(self):
        try:
            unread_messages_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '[href="https://www.linkedin.com/messaging/?"]')))
            try:
                unread_messages = unread_messages_element.find_element(By.CLASS_NAME, 'ember-view')
                unread_messages_show_element = unread_messages.find_element(By.CLASS_NAME, 'notification-badge--show')
                unread_messages_count_element = unread_messages_show_element.find_element(By.CLASS_NAME, 'notification-badge__count')
                unread_messages_count = unread_messages_count_element.text
            except Exception as e:
                unread_messages_count = 0

            notifications_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '[href="https://www.linkedin.com/notifications/?"]')))
            try:
                notifications_badge = notifications_element.find_element(By.CLASS_NAME, 'ember-view')
                notifications_show = notifications_badge.find_element(By.CLASS_NAME, 'notification-badge--show')
                notifications =  notifications_show.find_element(By.CLASS_NAME, 'notification-badge__count')
                notifications_count = notifications.text
            except Exception as e:
                notifications_count = 0

            return int(unread_messages_count), int(notifications_count)

        except Exception as e:
            print("An error occurred while retrieving unread counts:", e)
            return 0, 0

    def quit(self):
        if self.driver:
            self.driver.quit()


class LinkedInUnreadMonitorEmailer:
    def __init__(self, sender_email, sender_password, recipient_email, smtp_server, smtp_port):
        self.sender_email = sender_email
        self.sender_password = sender_password
        self.recipient_email = recipient_email
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port

    def send_email(self, subject, body):
        # Create a multipart message object
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = self.sender_email
        message["To"] = self.recipient_email

        # Create HTML content for the email
        html_content = body

        # Attach the HTML content to the message
        message.attach(MIMEText(html_content, "html"))

        # Setup the SMTP server and send the email
        with smtplib.SMTP(self.smtp_server,self.smtp_port) as server:
            server.starttls()
            server.login(self.sender_email, self.sender_password)
            server.sendmail(self.sender_email, self.recipient_email, message.as_string())


def generate_email_body(current_data, previous_data):
    css_style = '''
        body {
                    font-family: Arial, sans-serif;
                }
                
                table {
                    border-collapse: collapse;
                }
                
                th, td {
                    padding: 8px;
                    border: 1px solid black;
                }
                
                th {
                    background-color: #0077B5;
                    color: white;
                }
        '''
    body = f'''
    <html>
    <head>
        <style>
            {css_style}
        </style>
    </head>
    <body>
        <h1>LinkedIn Unread Messages and Notifications</h1>
        <table>
            <tr>
                <th>Metrics</th>
                <th>Previous</th>
                <th>Current</th>
                <th>Comparison</th>
            </tr>
            <tr>
                <td>Unread Messages</td>
                <td>{previous_data['unread_messages']}</td>
                <td>{current_data['unread_messages']}</td>
                <td>{current_data['unread_messages'] - previous_data['unread_messages']}</td>
            </tr>
            <tr>
                <td>Unread Notifications</td>
                <td>{previous_data['unread_notifications']}</td>
                <td>{current_data['unread_notifications']}</td>
                <td>{current_data['unread_notifications'] - previous_data['unread_notifications']}</td>
            </tr>
        </table>
    </body>
    </html>
    '''

    return body


def update_excel_data(filename, data):
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.append(list(data.values()))
    workbook.save(filename)


def retrieve_previous_data(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    return {
        "unread_messages": int(sheet["C"][-1].value),
        "unread_notifications": int(sheet["D"][-1].value)
    }


def main():
    username = 'gamocode@gmail.com'
    password = 'Swarup@123'
    chromedriver_path = '/path/to/chromedriver'
    sender_email = 'gamocode@gmail.com'
    sender_password = 'fmtdyfctgqgvomxb'
    recipient_email = 'gamodemy1@gmail.com'
    smtp_server = 'smtp.gmail.com'
    smtp_port = 587

    linkedin_monitor = LinkedInUnreadMonitor(chromedriver_path)
    linkedin_monitor.login(username, password)
    current_data = {
        'unread_messages': 0,
        'unread_notifications': 0
    }

    while True:
        workbook_filename = 'data.xlsx'
        # Retrieve the previous data from the Excel file
        previous_data = retrieve_previous_data(workbook_filename)
        # Update the Excel file with the current data
        unread_messages_count, notifications_count = linkedin_monitor.get_unread_counts()
        current_data['unread_messages'] = int(unread_messages_count)
        current_data['unread_notifications'] = int(notifications_count)
        update_excel_data(workbook_filename, {
            'Username': username,
            'Time': datetime.now().strftime("%d-%B, %H:%M"),
            'Unread Messages': current_data['unread_messages'],
            'Unread Notifications': current_data['unread_notifications'],
            'Messages Change': current_data['unread_messages'] - previous_data['unread_messages'],
            'Notifications Change': current_data['unread_notifications'] - previous_data['unread_notifications']
        })

        email_body = generate_email_body(current_data, previous_data)

        emailer = LinkedInUnreadMonitorEmailer(sender_email, sender_password, recipient_email, smtp_server, smtp_port)
        emailer.send_email('LinkedIn Unread Messages and Notifications', email_body)

        linkedin_monitor.quit()
        time.sleep(3 * 60 * 60)  # Sleep for 3 hours


if __name__ == '__main__':
    main()
